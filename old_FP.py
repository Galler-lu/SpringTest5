import time
import matplotlib.pyplot as plt
import openpyxl
class treeNode:
    def __init__(self, nameValue, numOccur, parentNode):
        self.name = nameValue
        self.count = numOccur
        self.nodeLink = None
        self.parent = parentNode
        self.children = {}

    def inc(self, numOccur):
        self.count += numOccur

    def disp(self, ind=1):
        print('  ' * ind, self.name, ' ', self.count)
        for child in self.children.values():
            child.disp(ind + 1)


def updateHeader(nodeToTest, targetNode):
    while nodeToTest.nodeLink != None:
        nodeToTest = nodeToTest.nodeLink
    nodeToTest.nodeLink = targetNode


def updateFPtree(items, inTree, headerTable, count):
    if items[0] in inTree.children:
        # 判断items的第一个结点是否已作为子结点
        inTree.children[items[0]].inc(count)
    else:
        # 创建新的分支
        inTree.children[items[0]] = treeNode(items[0], count, inTree)
        # 更新相应频繁项集的链表，往后添加
        if headerTable[items[0]][1] == None:
            headerTable[items[0]][1] = inTree.children[items[0]]
        else:
            updateHeader(headerTable[items[0]][1], inTree.children[items[0]])
    # 递归
    if len(items) > 1:
        updateFPtree(items[1::], inTree.children[items[0]], headerTable, count)


def createFPtree(dataSet, minSup=1):
    headerTable = {}
    for trans in dataSet:
        for item in trans:
            headerTable[item] = headerTable.get(item, 0) + dataSet[trans]
    for k in list(headerTable.keys()):
        if headerTable[k] < minSup:
            del (headerTable[k])  # 删除不满足最小支持度的元素

    print(headerTable)

    freqItemSet = set(headerTable.keys())  # 满足最小支持度的频繁项集
    if len(freqItemSet) == 0:
        return None, None
    for k in headerTable:
        headerTable[k] = [headerTable[k], None]  # element: [count, node]

    retTree = treeNode('Null Set', 1, None)
    for tranSet, count in dataSet.items():
        # dataSet：[element, count]
        localD = {}
        for item in tranSet:
            if item in freqItemSet:  # 过滤，只取该样本中满足最小支持度的频繁项
                localD[item] = headerTable[item][0]  # element : count
        if len(localD) > 0:
            # 根据全局频数从大到小对单样本排序
            orderedItem = [v[0] for v in sorted(localD.items(), key=lambda p: p[1], reverse=True)]
            # 用过滤且排序后的样本更新树
            updateFPtree(orderedItem, retTree, headerTable, count)
    return retTree, headerTable



# 构造成 element : count 的形式
def createInitSet(dataSet):
    retDict = {}

    for trans in dataSet:
        if frozenset(trans) in retDict.keys():
            retDict[frozenset(trans)] += 1
        else:
            retDict[frozenset(trans)] = 1
    return retDict



# 递归回溯
def ascendFPtree(leafNode, prefixPath):
    if leafNode.parent != None:
        prefixPath.append(leafNode.name)
        ascendFPtree(leafNode.parent, prefixPath)


# 条件模式基
def findPrefixPath(basePat, myHeaderTab):
    treeNode = myHeaderTab[basePat][1]  # basePat在FP树中的第一个结点
    condPats = {}
    while treeNode != None:
        prefixPath = []
        ascendFPtree(treeNode, prefixPath)  # prefixPath是倒过来的，从treeNode开始到根
        if len(prefixPath) > 1:
            condPats[frozenset(prefixPath[1:])] = treeNode.count  # 关联treeNode的计数
        treeNode = treeNode.nodeLink  # 下一个basePat结点
    return condPats


def mineFPtree(inTree, headerTable, minSup, preFix, freqItemList):
    # 最开始的频繁项集是headerTable中的各元素
    # bigL = [v[0] for v in sorted(headerTable.items(), key=lambda p: p[1])]  # 根据频繁项的总频次排序
    bigL = [v[0] for v in sorted(headerTable.items(), key=lambda p: p[1][0])]
    for basePat in bigL:  # 对每个频繁项
        newFreqSet = preFix.copy()
        newFreqSet.add(basePat)
        freqItemList.append(newFreqSet)
        condPattBases = findPrefixPath(basePat, headerTable)  # 当前频繁项集的条件模式基
        myCondTree, myHead = createFPtree(condPattBases, minSup)  # 构造当前频繁项的条件FP树
        if myHead != None:
            mineFPtree(myCondTree, myHead, minSup, newFreqSet, freqItemList)  # 递归挖掘条件FP树

#读取数据集
def read_excel(fileName):
    my_list = []
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    maxrows = ws.max_row
    for i in range(maxrows - 1):
        temp_list = []
        for each in ws.iter_cols(min_row=2,min_col=2):
            str=each[i].value.removeprefix('{')
            temp_list.append(str.removesuffix('}').split(','))
        my_list.extend(temp_list)
    return my_list
if __name__ == '__main__':

    simpDat = read_excel("test.xlsx")
    plt.figure()
    frequently_list = []
    count_list = []
    # 调参：可以调置信度或组合数
    params = range(10, 13, 1)

    for i in params:
        initSet = createInitSet(simpDat)
        count_start_time = time.time()
        myFPtree, myHeaderTab = createFPtree(initSet, minSup=i)
        # myFPtree.disp()
        start_time = time.time()
        freqItems = []
        mineFPtree(myFPtree, myHeaderTab, i, set([]), freqItems)
        end_time = time.time()
        count_end_time = time.time()
        frequently_list.append(end_time - start_time)
        count_list.append(count_end_time - count_start_time)
    plt.subplot(1, 2, 1)
    plt.plot(params, frequently_list)
    plt.xlabel("Confidence")
    plt.xticks(params)
    plt.ylabel("Running time of frequent itemsets /s")
    plt.subplot(1, 2, 2)
    plt.plot(params, count_list)
    plt.xlabel("Support/time")
    plt.xticks(params)
    plt.ylabel("Total operating time/s")
    plt.show()