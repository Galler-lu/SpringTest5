import math
import time

import openpyxl

import matplotlib.pyplot as plt


class treeNode:
    def __init__(self, nameValue, numOccur, parentNode):
        self.name = nameValue  # 节点名称
        self.count = numOccur  # 节点出现的次数
        self.nodeLink = None  # 链接指向的下一个节点
        self.parent = parentNode  # 父节点
        self.children = {}  # 子节点

    def inc(self, numOccur):  # 该函数用来增加节点出现的次数
        self.count += numOccur

    def disp(self, ind=1):
        print('  ' * ind, self.name, ' ', self.count)
        for child in self.children.values():
            child.disp(ind + 1)  # 展示节点名称和出现的次数


from collections import OrderedDict


def createInitSet(dataSet):
    retDict = {}
    for trans in dataSet:
        if frozenset(trans) in retDict.keys():
            retDict[frozenset(trans)] += 1
        else:
            retDict[frozenset(trans)] = 1
    return retDict


def createTree(dataSet, minSup=1, num=1):
    headerTable = {}  # 用来存储每项元素及其出现次数
    for trans in dataSet:  # 遍历每条记录
        for item in trans:  # 遍历每条记录的每项元素
            headerTable[item] = headerTable.get(item, 0) + dataSet[trans]  # 计算每项元素的出现次数
    for k in list(headerTable.keys()):
        if headerTable[k] < minSup:
            del (headerTable[k])  # 如果某项元素的支持度小于最小支持度，从headerTable中删掉该元素
    print(headerTable)
    # 1、自己写的代码#####################################################################################
    test = sorted(headerTable.items(), key=lambda item: item[1], reverse=True)
    # 取其前三
    keyList = []
    valueList = []
    for i in range(num):
        keyList.append(test[:num:][i][0])
        valueList.append(test[:num:][i][1])
    import itertools
    keyList_combination = list(itertools.combinations(keyList, 2))
    new_dict = {}
    for item in keyList_combination:
        new_dict[item] = 0
    log_list = []
    for item in new_dict:
        log_list.append(headerTable[item[-1]])
    log_list_new_dict = dict(zip(keyList_combination, log_list))
    # 1、自己写的代码######################################################################################

    freqItemSet = set(headerTable.keys())  # freqItemSet中的每一项元素的支持度均大于或等于最小支持度

    if len(freqItemSet) == 0:
        return None, None
    for k in headerTable:
        headerTable[k] = [headerTable[k], None]
    retTree = treeNode('Null Set', 1, None)  # 创建根节点
    for tranSet, count in dataSet.items():  # 遍历每一条事务数据

        # 2、自己写的代码######################################################################################
        # print("tranSet的类型为",type(tranSet))
        flag_tranSet = list(tranSet)
        # print("flag_tranSet为", flag_tranSet)
        for i in list(keyList_combination):
            if (set(i) < set(flag_tranSet)):
                new_dict[i] += 1
    import operator
    list1 = []
    for item in new_dict:
        if new_dict[item] >= math.ceil(log_list_new_dict[item] / 2):
            # 可以合并的记录为-1
            new_dict[item] = -1
            list1.append(item)

    # 2、自己写的代码################################################################################################

    for tranSet, count in dataSet.items():  # 遍历每一条事务数据

        localD = {}
        for item in tranSet:  # 遍历这条数据中的每个元素
            if item in freqItemSet:  # 过滤每条记录中支持度小于最小支持度的元素
                localD[item] = headerTable[item][0]  # 把headerTable中记录的该元素的出现次数赋值给localD中的对应键
        if len(localD) > 0:  # 如果该条记录有符合条件的元素
            orderedItems = [v[0] for v in
                            sorted(localD.items(), key=lambda p: p[1], reverse=True)]  # 元素按照支持度排序，支持度越大，排位越靠前
            # 3、###############################################################################################
            oder_Flag = []
            j = 0
            while (j <= len(orderedItems) - 1):
                list2 = []
                tuple1 = tuple(orderedItems[j:j + 2])
                list2.append(tuple1)

                if (all(elem in list1 for elem in list2)):

                    orderedItem_1 = "".join(list(tuple1))
                    headerTable[orderedItem_1] = [-2, None]
                    j += 2
                    oder_Flag.append(orderedItem_1)
                    list2.clear()
                else:
                    oder_Flag.extend(list(tuple(orderedItems[j:j + 2])))
                    j += 1

            oder_Flag = sorted(set(oder_Flag), key=oder_Flag.index)

            updateTree(oder_Flag, retTree, headerTable, count)
    return retTree, headerTable


def updateTree(items, inTree, headerTable, count):
    if items[0] in inTree.children:  # 如果inTree的子节点中已经存在该元素
        inTree.children[items[0]].inc(count)  # 树中该元素增加值，增加的值为该元素所在记录的出现次数
    else:
        inTree.children[items[0]] = treeNode(items[0], count, inTree)  # 如果树中不存在该元素，重新创建一个节点
        if headerTable[items[0]][1] == None:  # 如果在相似元素的字典headerTable中，该元素键对应的列表值中，起始元素为None
            headerTable[items[0]][1] = inTree.children[items[0]]  # 把新创建的这个节点赋值给起始元素
        else:
            updateHeader(headerTable[items[0]][1],
                         inTree.children[items[0]])  # 如果在相似元素字典headerTable中，该元素键对应的值列表中已经有了起始元素，那么把这个新建的节点放到值列表的最后
    if len(items) > 1:  # 如果在这条记录中，符合条件的元素个数大于1
        updateTree(items[1::], inTree.children[items[0]], headerTable, count)  # 从第二个元素开始，递归调用updateTree函数。


def updateHeader(nodeToTest, targetNode):  # 该函数实现把targetNode放到链接的末端
    while (nodeToTest.nodeLink != None):
        nodeToTest = nodeToTest.nodeLink
    nodeToTest.nodeLink = targetNode


def ascendTree(leafNode, prefixPath):  # 该函数找出元素节点leafNode的所有前缀路径，并把包括该leafNode及其前缀路径的各个节点的名称保存在prefixPath中
    if leafNode.parent != None:
        prefixPath.append(leafNode.name)
        ascendTree(leafNode.parent, prefixPath)


def findPrefixPath(basePat, myHeaderTab):
    treeNode = myHeaderTab[basePat][1]  # basePat在FP树中的第一个结点
    condPats = {}
    while treeNode != None:
        prefixPath = []
        ascendTree(treeNode, prefixPath)  # prefixPath是倒过来的，从treeNode开始到根
        if len(prefixPath) > 1:
            condPats[frozenset(prefixPath[1:])] = treeNode.count  # 关联treeNode的计数
        treeNode = treeNode.nodeLink  # 下一个basePat结点
    return condPats


def mineFPtree(inTree, headerTable, minSup, preFix, freqItemList):
    bigL = [v[0] for v in sorted(headerTable.items(), key=lambda p: p[1][0])]
    for basePat in bigL:  # 对每个频繁项
        newFreqSet = preFix.copy()
        newFreqSet.add(basePat)
        freqItemList.append(newFreqSet)

        condPattBases = findPrefixPath(basePat, headerTable)  # 当前频繁项集的条件模式基
        myCondTree, myHead = createFPtree1(condPattBases, minSup)  # 构造当前频繁项的条件FP树
        if myHead != None:
            mineFPtree(myCondTree, myHead, minSup, newFreqSet, freqItemList)  # 递归挖掘条件FP树


def createFPtree1(dataSet, minSup=1):
    headerTable = {}
    for trans in dataSet:
        for item in trans:
            headerTable[item] = headerTable.get(item, 0) + dataSet[trans]
    for k in list(headerTable.keys()):
        if headerTable[k] < minSup:
            del (headerTable[k])  # 删除不满足最小支持度的元素
    freqItemSet = set(headerTable.keys())  # 满足最小支持度的频繁项集
    if len(freqItemSet) == 0:
        return None, None
    for k in headerTable:
        headerTable[k] = [headerTable[k], None]  # element: [count, node]

    retTree = treeNode('Null Set', 1, None)
    for tranSet, count in dataSet.items():
        # dataSet：[element, count]
        localD = {}
        for item in tranSet:
            if item in freqItemSet:  # 过滤，只取该样本中满足最小支持度的频繁项
                localD[item] = headerTable[item][0]  # element : count
        if len(localD) > 0:
            # 根据全局频数从大到小对单样本排序
            orderedItem = [v[0] for v in sorted(localD.items(), key=lambda p: p[1], reverse=True)]
            # 用过滤且排序后的样本更新树
            updateTree(orderedItem, retTree, headerTable, count)
    return retTree, headerTable


# 读取数据集
def read_excel(fileName):
    my_list = []
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    maxrows = ws.max_row
    for i in range(maxrows - 1):
        temp_list = []
        for each in ws.iter_cols(min_row=2, min_col=2):
            str = each[i].value.removeprefix('{')
            temp_list.append(str.removesuffix('}').split(','))
        my_list.extend(temp_list)
    return my_list


if __name__ == '__main__':

    simpDat = read_excel("test.xlsx")
    initSet = createInitSet(simpDat)
    plt.figure()
    frequently_list = []
    count_list = []
    # 调参：可以调置信度或组合数
    params = range(10, 15, 1)
    for i in params:
        count_start_time = time.time()
        myFPtree, myHeaderTab = createTree(initSet, 10, i)
        # myFPtree.disp()
        start_time = time.time()
        freqItems = []
        mineFPtree(myFPtree, myHeaderTab, 10, set([]), freqItems)
        end_time = time.time()
        count_end_time = time.time()
        frequently_list.append(end_time - start_time)
        count_list.append(count_end_time - count_start_time)
    plt.subplot(1, 2, 1)
    plt.plot(params, frequently_list)
    plt.xlabel("Confidence")
    plt.xticks(params)
    plt.ylabel("Running time of frequent itemsets /s")
    plt.subplot(1, 2, 2)
    plt.plot(params, count_list)
    plt.xlabel("Support/time")
    plt.xticks(params)
    plt.ylabel("Total operating time/s")
    plt.show()
